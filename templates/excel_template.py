"""
Excel template styles – matches demo file exactly.
Colour palette, fonts, fills, borders, and helper utilities.
"""

from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ───────────────────────────────────────────────────────────
# Colour Palette (ARGB – demo exact hex)
# ───────────────────────────────────────────────────────────
COLORS = {
    # Primary blues
    "dark_navy":      "FF1F3864",
    "medium_blue":    "FF2F5597",
    "light_blue_bg":  "FFBDD7EE",

    # Greens
    "dark_green":     "FF375623",
    "light_green_bg": "FFE2EFDA",

    # Oranges
    "orange":         "FFED7D31",
    "gold":           "FFC9A84C",
    "light_orange_bg":"FFFCE9DA",

    # Yellows
    "light_yellow_bg":"FFFFF2CC",

    # Reds
    "dark_red":       "FFC00000",
    "light_red_bg":   "FFFCE4D6",

    # Purples
    "purple":         "FF7030A0",
    "light_purple_bg":"FFEAD1F5",

    # Neutrals
    "white":          "FFFFFFFF",
    "off_white":      "FFF2F2F2",
    "grey_text":      "FF595959",
    "black":          "FF000000",

    # Benchmark colours
    "bench_green":    "FF6AA84F",
    "bench_blue":     "FF3C78D8",
    "bench_gold":     "FFF1C232",
    "input_blue":     "FF0000FF",

    # Upgrade Path – AirtoniQ exact hex
    "dark_header_navy":"FF1F4E78",
    "header_blue":     "FF2E5090",
    "orange_tier":     "FFC65911",
    "green_tier":      "FF548235",
    "yellow_header":   "FFFFF2CC",
    "orange_light_bg": "FFF4B084",
    "green_light_bg2": "FFA8D08D",
    "lavender_bg":     "FFD9D2E9",
    "grey_light_bg":   "FFD6DCE4",
}

# ───────────────────────────────────────────────────────────
# Fills
# ───────────────────────────────────────────────────────────
def solid_fill(color_key: str) -> PatternFill:
    return PatternFill(start_color=COLORS[color_key], end_color=COLORS[color_key], fill_type="solid")

FILLS = {k: solid_fill(k) for k in COLORS}

# ───────────────────────────────────────────────────────────
# Fonts
# ───────────────────────────────────────────────────────────
def make_font(size=9, bold=False, color_key="black"):
    return Font(name="Calibri", size=size, bold=bold, color=COLORS[color_key])

FONTS = {
    # Sheet titles
    "title_18_white":    make_font(18, True, "white"),
    "title_14_white":    make_font(14, True, "white"),
    "subtitle_11_gold":  make_font(11, False, "gold"),
    "section_12_white":  make_font(12, True, "white"),
    "subsection_10_white": make_font(10, True, "white"),
    "subsection_10_navy":  make_font(10, True, "dark_navy"),

    # Body text
    "body_9":            make_font(9),
    "body_9_white":      make_font(9, False, "white"),
    "body_9_bold":       make_font(9, True),
    "body_9_bold_navy":  make_font(9, True, "dark_navy"),
    "body_9_bold_orange":make_font(9, True, "orange"),
    "body_9_bold_blue":  make_font(9, True, "medium_blue"),
    "body_9_bold_green": make_font(9, True, "dark_green"),
    "body_9_bold_purple":make_font(9, True, "purple"),
    "body_9_bold_red":   make_font(9, True, "dark_red"),
    "body_9_grey":       make_font(9, False, "grey_text"),
    "body_9_green":      make_font(9, False, "dark_green"),
    "body_9_navy":       make_font(9, False, "dark_navy"),
    "body_9_orange":     make_font(9, False, "orange"),
    "body_9_purple":     make_font(9, False, "purple"),
    "body_9_red":        make_font(9, False, "dark_red"),
    "body_9_blue":       make_font(9, False, "medium_blue"),
    "body_10_white":     make_font(10, False, "white"),
    "body_10_bold_blue": make_font(10, True, "medium_blue"),
    "body_10_bold_orange": make_font(10, True, "orange"),
    "body_10_bold_green":  make_font(10, True, "dark_green"),
    "body_11_white":     make_font(11, False, "white"),
    "body_11_blue":      make_font(11, False, "medium_blue"),
    "body_11_orange":    make_font(11, False, "orange"),
    "body_11_green":     make_font(11, False, "dark_green"),
    "body_11_grey":      make_font(11, False, "grey_text"),
    "body_11_bold_navy": make_font(11, True, "dark_navy"),
    "body_12_bold_purple": make_font(12, True, "purple"),

    # Small text
    "small_8":           make_font(8),
    "small_8_white":     make_font(8, True, "white"),
    "small_8_grey":      make_font(8, False, "grey_text"),
    "small_8_bold":      make_font(8, True),
    "small_8_bold_navy": make_font(8, True, "dark_navy"),
    "small_8_bold_purple": make_font(8, True, "purple"),

    # Checkbox
    "checkbox_11_navy":  make_font(11, True, "dark_navy"),
    "checkbox_11_purple":make_font(11, True, "purple"),
    "checkbox_11_grey":  make_font(11, True, "grey_text"),

    # Benchmark
    "bench_green":       make_font(9, False, "bench_green"),
    "bench_blue":        make_font(9, True, "bench_blue"),
    "bench_gold":        make_font(9, True, "bench_gold"),
    "input_blue":        make_font(9, False, "input_blue"),

    # Upgrade Path
    "body_10_bold_dark_navy": make_font(10, True, "dark_header_navy"),
    "body_10":           make_font(10),
    "body_12_bold_white": make_font(12, True, "white"),
    "body_13_bold_white": make_font(13, True, "white"),
    "body_12_bold_navy":  make_font(12, True, "dark_header_navy"),
    "body_9_italic":      Font(name="Calibri", size=9, italic=True),
}

# ───────────────────────────────────────────────────────────
# Borders
# ───────────────────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin", color="FFD9D9D9"),
    right=Side(style="thin", color="FFD9D9D9"),
    top=Side(style="thin", color="FFD9D9D9"),
    bottom=Side(style="thin", color="FFD9D9D9"),
)

NO_BORDER = Border()

# ───────────────────────────────────────────────────────────
# Alignment
# ───────────────────────────────────────────────────────────
ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ───────────────────────────────────────────────────────────
# Helper Functions
# ───────────────────────────────────────────────────────────

def style_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    """Set value and style for a single cell."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    else:
        cell.alignment = ALIGN_LEFT
    if border:
        cell.border = border
    return cell


def style_range(ws, row, col_start, col_end, value, font=None, fill=None, alignment=None, merge=True):
    """Write value in first cell and optionally merge across columns."""
    cell = style_cell(ws, row, col_start, value, font, fill, alignment)
    if merge and col_end > col_start:
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row, end_column=col_end,
        )
        # Apply fill to all cells in merged range
        for c in range(col_start, col_end + 1):
            ws.cell(row=row, column=c).fill = fill or PatternFill()
            ws.cell(row=row, column=c).alignment = alignment or ALIGN_LEFT
    return cell


def style_merged_block(ws, row_start, row_end, col_start, col_end, value, font=None, fill=None, alignment=None):
    """Merge a block of cells (multiple rows and columns) and style them."""
    ws.merge_cells(
        start_row=row_start, start_column=col_start,
        end_row=row_end, end_column=col_end,
    )
    cell = ws.cell(row=row_start, column=col_start, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    else:
        cell.alignment = ALIGN_LEFT
    # Apply fill to entire block
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            ws.cell(row=r, column=c).fill = fill or PatternFill()
    return cell


def set_column_widths(ws, widths: dict):
    """Set column widths from a dict like {'A': 3, 'B': 40, ...}."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def alternating_row_fill(row_index: int):
    """Return off_white for even rows, white for odd (0-indexed from data start)."""
    return FILLS["off_white"] if row_index % 2 == 0 else FILLS["white"]


def write_lock_footer(ws, row_start, row_end, col_start, col_end, text):
    """Write the 🔒 upgrade teaser footer at the bottom of a sheet."""
    style_merged_block(
        ws, row_start, row_end, col_start, col_end,
        text,
        font=FONTS["body_9_white"],
        fill=FILLS["dark_navy"],
        alignment=ALIGN_LEFT,
    )
