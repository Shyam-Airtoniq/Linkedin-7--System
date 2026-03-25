"""Sheet builders for Sheets 9-10 of the LinkedIn COMPLETE SYSTEM."""
import os
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font
from templates.excel_template import (
    COLORS, FILLS, FONTS, ALIGN_LEFT, ALIGN_CENTER, ALIGN_LEFT_CENTER,
    style_cell, style_range, style_merged_block, set_column_widths,
)


def build_sheet9_results_tracker(wb, tracker_data):
    """Sheet 9: Advanced Results Tracker — with A/B testing and benchmarks."""
    ws = wb.create_sheet("9. Results Tracker")
    set_column_widths(ws, {"A": 3, "B": 32, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16})

    style_range(ws, 1, 2, 7,
        "ADVANCED RESULTS TRACKER  |  Complete System",
        FONTS["title_14_white"], FILLS["dark_navy"], ALIGN_LEFT_CENTER)

    # Benchmarks
    style_range(ws, 3, 2, 7, "📊  KEY PERFORMANCE BENCHMARKS",
        FONTS["section_12_white"], FILLS["dark_navy"], ALIGN_LEFT_CENTER)
    r = 4
    for col, hdr in [(2, "Metric"), (4, "Good"), (5, "Great"), (6, "Exceptional")]:
        style_cell(ws, r, col, hdr, FONTS["body_9_white"], FILLS["medium_blue"], ALIGN_CENTER)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    r += 1
    for i, b in enumerate(tracker_data.get("benchmarks", [])):
        fill = FILLS["light_blue_bg"] if i % 2 == 0 else FILLS["white"]
        style_cell(ws, r, 2, b.get("metric", ""), FONTS["body_9"], fill, ALIGN_LEFT)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        style_cell(ws, r, 4, b.get("good", ""), FONTS["body_9"], fill, ALIGN_CENTER)
        style_cell(ws, r, 5, b.get("great", ""), FONTS["body_9_bold_orange"], fill, ALIGN_CENTER)
        style_cell(ws, r, 6, b.get("exceptional", ""), FONTS["body_9_bold_green"], fill, ALIGN_CENTER)
        for c in range(2, 7):
            ws.cell(row=r, column=c).fill = fill
        r += 1

    # Weekly Tracker template
    r += 1
    style_range(ws, r, 2, 7, "📈  WEEKLY PERFORMANCE TRACKER",
        FONTS["section_12_white"], FILLS["medium_blue"], ALIGN_LEFT_CENTER)
    r += 1
    weeks = ["Metric", "Week 1", "Week 2", "Week 3", "Week 4"]
    for j, w in enumerate(weeks):
        style_cell(ws, r, 2 + j, w, FONTS["body_9_white"], FILLS["dark_navy"], ALIGN_CENTER)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
    r += 1
    metric_names = ["Profile Views", "Post Impressions", "Engagement Rate",
                    "Connections Sent", "Acceptance Rate", "DMs Sent",
                    "DM Reply Rate", "Calls Booked", "Content Published"]
    for i, m in enumerate(metric_names):
        fill = FILLS["off_white"] if i % 2 == 0 else FILLS["white"]
        style_cell(ws, r, 2, m, FONTS["body_9"], fill, ALIGN_LEFT)
        for c in range(3, 7):
            style_cell(ws, r, c, "[ Enter ]", FONTS["small_8_grey"], fill, ALIGN_CENTER)
        r += 1

    # A/B Testing Framework
    r += 1
    style_range(ws, r, 2, 7, "🧪  A/B TESTING FRAMEWORK",
        FONTS["section_12_white"], FILLS["purple"], ALIGN_LEFT_CENTER)
    r += 1
    ab_headers = ["Test Name", "Variable A", "Variable B", "Metric", "Duration", "How to Measure"]
    for j, h in enumerate(ab_headers):
        style_cell(ws, r, 2 + j, h, FONTS["small_8_white"], FILLS["dark_navy"], ALIGN_LEFT)
    r += 1
    for i, ab in enumerate(tracker_data.get("ab_testing_framework", [])):
        fill = FILLS["light_purple_bg"] if i % 2 == 0 else FILLS["white"]
        vals = [ab.get("test_name",""), ab.get("variable_a",""), ab.get("variable_b",""),
                ab.get("metric_to_track",""), ab.get("duration",""), ab.get("how_to_measure","")]
        for j, v in enumerate(vals):
            style_cell(ws, r, 2 + j, v, FONTS["small_8"], fill, ALIGN_LEFT)
        r += 1

    # Testing Calendar
    r += 1
    style_range(ws, r, 2, 7, "📅  12-WEEK TESTING CALENDAR",
        FONTS["section_12_white"], FILLS["dark_green"], ALIGN_LEFT_CENTER)
    r += 1
    for i, tc in enumerate(tracker_data.get("testing_calendar", [])):
        fill = FILLS["light_green_bg"] if i % 2 == 0 else FILLS["white"]
        style_cell(ws, r, 2, tc.get("week", ""), FONTS["body_9_bold_navy"], fill, ALIGN_LEFT)
        style_cell(ws, r, 3, tc.get("test", ""), FONTS["body_9_bold_green"], fill, ALIGN_LEFT)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        style_range(ws, r, 5, 7, tc.get("action", ""), FONTS["body_9"], fill, ALIGN_LEFT)
        r += 1

    # Benchmark Comparison
    r += 1
    bc = tracker_data.get("benchmark_comparison", {})
    style_range(ws, r, 2, 7, "🏆  YOUR RESULTS vs. INDUSTRY AVERAGES",
        FONTS["section_12_white"], FILLS["orange"], ALIGN_LEFT_CENTER)
    r += 1
    for col, hdr in [(2, "Metric"), (3, "Industry Average"), (5, "Your Target"), (6, "Multiplier")]:
        style_cell(ws, r, col, hdr, FONTS["small_8_white"], FILLS["dark_navy"], ALIGN_CENTER)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    r += 1
    for i, ia in enumerate(bc.get("industry_avg", [])):
        fill = FILLS["light_orange_bg"] if i % 2 == 0 else FILLS["white"]
        style_cell(ws, r, 2, ia.get("metric", ""), FONTS["body_9"], fill, ALIGN_LEFT)
        style_cell(ws, r, 3, ia.get("industry_avg", ""), FONTS["body_9"], fill, ALIGN_CENTER)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        style_cell(ws, r, 5, ia.get("your_target", ""), FONTS["body_9_bold_green"], fill, ALIGN_CENTER)
        style_cell(ws, r, 6, ia.get("multiplier", ""), FONTS["body_9_bold_orange"], fill, ALIGN_CENTER)
        r += 1


def build_sheet10_upgrade(wb, logo_path=None):
    """Sheet 10: Upgrade Path (STATIC — same for everyone).
    Points users to the Done-With-You ($1,999) tier.
    """
    ws = wb.create_sheet("10. Upgrade Path")
    set_column_widths(ws, {"A": 3, "B": 48, "C": 30, "D": 30, "E": 30, "F": 8.71})

    F_NAVY  = FILLS["dark_header_navy"]
    F_HBLUE = FILLS["header_blue"]
    F_ORANG = FILLS["orange_tier"]
    F_GREEN = FILLS["green_tier"]
    F_YEL   = FILLS["yellow_header"]
    F_ORBG  = FILLS["orange_light_bg"]
    F_GRBG  = FILLS["green_light_bg2"]
    F_LAV   = FILLS["lavender_bg"]
    F_GREY  = FILLS["grey_light_bg"]
    F_WHITE = FILLS["white"]

    FONT_TITLE   = FONTS["title_14_white"]
    FONT_HDR11W  = Font(name="Calibri", size=11, bold=True, color=COLORS["white"])
    FONT_HDR12W  = FONTS["body_12_bold_white"]
    FONT_HDR13W  = FONTS["body_13_bold_white"]
    FONT_HDR12N  = FONTS["body_12_bold_navy"]
    FONT_B10     = FONTS["body_10"]
    FONT_B10_BN  = FONTS["body_10_bold_dark_navy"]
    FONT_B9      = FONTS["body_9"]
    FONT_B9_IT   = FONTS["body_9_italic"]

    # ─── Row 1: Title ───
    style_range(ws, 1, 2, 5,
        "NEXT LEVEL  |  Complete System vs Done-With-You",
        FONT_TITLE, F_NAVY, ALIGN_LEFT_CENTER)
    ws.row_dimensions[1].height = 30

    # ─── Row 3: Tier headers ───
    style_cell(ws, 3, 2, "FEATURE", FONT_HDR11W, F_NAVY, ALIGN_LEFT_CENTER)
    style_cell(ws, 3, 3, "🚀 COMPLETE SYSTEM (You Have This)", FONT_HDR11W, F_ORANG, ALIGN_CENTER)
    style_cell(ws, 3, 4, "💎 DONE-WITH-YOU", FONT_HDR11W, F_GREEN, ALIGN_CENTER)
    ws.row_dimensions[3].height = 34.5

    # ─── Row 4: Pricing ───
    style_cell(ws, 4, 3, "$199 (Your Current Plan)", FONT_HDR12N, F_ORBG, ALIGN_CENTER)
    style_cell(ws, 4, 4, "$1,999 One-Time (90 Days)", FONT_HDR12N, F_GRBG, ALIGN_CENTER)
    ws.row_dimensions[4].height = 27.75

    # ─── Feature comparison (Complete vs DWY) ───
    FEATURES = [
        ("■ WHAT'S INCLUDED", [
            ("AI-Powered Strategic Analysis", "✅ Included", "✅ Customized by Our Team"),
            ("Profile Optimization", "✅ 25 Headlines + Story Arc", "✅ We Write & Optimize Your Profile"),
            ("Content Calendar", "✅ 90-Day Calendar (36 Posts)", "✅ First Month Written For You"),
            ("Content Hooks Library", "✅ 50 Proven Hooks", "✅ Custom Hook Library"),
            ("Story Templates", "✅ 15 Templates", "✅ Story Mining Session"),
            ("Outreach Templates", "✅ 8 CR + 6 DM Sequences", "✅ We Write Your Sequences"),
            ("Email Sequence", "✅ 5-Email Sequence", "✅ Multiple Sequences with A/B Testing"),
            ("Sales Navigator Guide", "✅ Setup + Optimization Guide", "✅ We Set Up Your Account"),
            ("Discovery Script", "✅ Complete Word-for-Word Script", "✅ Recorded Role-Play Session"),
            ("Objection Scripts", "✅ 15 Scripts", "✅ Objection Handling Training"),
            ("Meeting Templates", "✅ 3 Templates + Calendar Guide", "✅ We Set Up Your Calendly"),
            ("Lead Magnet", "✅ 10 Formats + 5 Landing Pages", "✅ We Create Your Lead Magnet"),
            ("Email Nurture", "✅ 7-Email Welcome Sequence", "✅ Sequences Written For You"),
            ("Implementation", "✅ Day-by-Day 90-Day Plan", "✅ Personalized Timeline"),
            ("Analytics", "✅ A/B Testing + Benchmarks", "✅ Monthly Optimization Reports"),
        ]),
        ("■ SUPPORT & ACCESS", [
            ("Email Support", "✅ 48-Hour Response", "✅ Priority 24-Hour Response"),
            ("Community Access", "✅ Skool Community", "✅ + Direct Access to Founder"),
            ("Strategy Sessions", "✅ One Session (45 Min)", "✅ 6 Sessions (Bi-Weekly)"),
            ("Weekly Group Q&A", "—", "✅ 12 Calls (90 Days)"),
            ("Done-For-You Deliverables", "—", "✅ Profile + 3 Posts + Lead Magnet"),
            ("Video Tutorials", "—", "✅ Private Training Sessions"),
        ]),
    ]

    r = 5
    for section_title, features in FEATURES:
        style_range(ws, r, 2, 4, section_title, FONT_HDR11W, F_HBLUE, ALIGN_LEFT_CENTER)
        ws.row_dimensions[r].height = 24.75
        r += 1
        for i, (feat, complete, dwy) in enumerate(features):
            ws.row_dimensions[r].height = 31.5
            feat_fill = F_LAV if i % 2 == 0 else F_WHITE
            style_cell(ws, r, 2, feat, FONT_B10, feat_fill, ALIGN_LEFT_CENTER)
            style_cell(ws, r, 3, complete, FONT_B9, F_ORBG, ALIGN_CENTER)
            style_cell(ws, r, 4, dwy, FONT_B9, F_GRBG, ALIGN_CENTER)
            r += 1

    # ─── ROI Section ───
    ws.row_dimensions[r].height = 15.75
    r += 1
    style_range(ws, r, 2, 4, "💰 WHAT DONE-WITH-YOU ADDS",
        FONT_HDR12W, F_NAVY, ALIGN_LEFT_CENTER)
    ws.row_dimensions[r].height = 24.75
    r += 1
    value_rows = [
        ("Your Investment", "$199 (Complete System)", "$1,999 (Done-With-You)"),
        ("Equivalent Market Value", "$800–$1,200", "$5,000–$10,000"),
        ("Time to First Lead", "14-21 days (Guided)", "7-14 days (Done-With-You)"),
        ("Expected 90-Day Result", "15-25 qualified leads", "30+ qualified leads"),
        ("Cost Per Lead", "$7–$13/lead", "$50–$65/lead"),
    ]
    for label, comp_val, dwy_val in value_rows:
        ws.row_dimensions[r].height = 21.75
        style_cell(ws, r, 2, label, FONT_B10_BN, F_WHITE, ALIGN_LEFT_CENTER)
        style_cell(ws, r, 3, comp_val, FONT_B9, F_ORBG, ALIGN_CENTER)
        style_cell(ws, r, 4, dwy_val, FONT_B9, F_GRBG, ALIGN_CENTER)
        r += 1

    # ─── CTA ───
    ws.row_dimensions[r].height = 15.75
    r += 1
    style_range(ws, r, 2, 4,
        "💎 READY FOR DONE-WITH-YOU?",
        FONT_HDR13W, F_NAVY, ALIGN_LEFT_CENTER)
    ws.row_dimensions[r].height = 30
    r += 1
    style_merged_block(ws, r, r + 2, 2, 4,
        "You already have the complete playbook. Done-With-You means we execute it WITH you.\n"
        "→  Book a strategy call: Let's review your Complete System results and plan next steps.\n"
        "→  Contact: shyam@9gravity.io  |  LinkedIn: /in/shyam9gravity",
        FONT_B10, F_GREY, ALIGN_LEFT)
